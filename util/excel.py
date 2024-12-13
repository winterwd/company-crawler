#!/usr/bin/python3
# -*-: coding: utf-8 -*-
"""
:author: winter
:date: 02/10/2022
:desc: excel生成工具类
"""

import os
from openpyxl import Workbook, load_workbook
from db.models import Company


class CellItem(object):
    def __init__(self, src: Company = None):
        if src is None:
            self.keyword = '关键词'
            self.short_name = '企业简称'
            self.name = '企业名称'
            self.company_area = '所属区域'
            self.company_address = '联系地址'
            self.contact = '联系方式'
            self.representative = '企业法人'
            self.biz_status = '经营状态'
            self.company_type = '企业类型'
            self.register_capital = '注册资金'
            self.actual_capital = '实缴资金'
            self.found_time = '成立时间'
            self.tags = '标签列表'
            self.industry = '行业分类'
        else:
            self.keyword = src.keyword
            self.short_name = src.short_name
            self.name = src.name
            self.company_area = f'{src.province}{src.city}{src.district}'
            self.company_address = src.company_address

            if src.phones is not None and len(src.phones) > 1:
                self.contact = ', '.join(src.phones)
            elif src.contact is not None:
                self.contact = src.contact
            else:
                self.contact = '暂无'

            self.representative = src.representative
            self.biz_status = src.biz_status
            self.found_time = src.found_time
            self.company_type = src.company_type
            self.register_capital = src.register_capital
            self.actual_capital = src.actual_capital
            if src.tags is None:
                self.tags = '无'
            else:
                self.tags = ','.join(src.tags)
            self.industry = src.industry

    def values(self):
        return [
            self.keyword,
            self.short_name,
            self.name,
            self.company_area,
            self.company_address,
            self.contact,
            self.representative,
            self.biz_status,
            self.company_type,
            self.register_capital,
            self.actual_capital,
            self.found_time,
            self.tags,
            self.industry,
        ]


def get_new_sheet_name(existing_sheets, base_name="Sheet"):
    """
    根据已有的工作表名称生成新的 sheet 名称。
    :param existing_sheets: 已存在的工作表名称列表
    :param base_name: 工作表的基础名称（默认是 "Sheet"）
    :return: 新的工作表名称
    """
    index = 0
    while True:
        new_name = f"{base_name}{index if index > 0 else ''}"  # "Sheet", "Sheet1", "Sheet2", ...
        if new_name not in existing_sheets:
            return new_name
        index += 1


def write(file_path: str, value: list[Company]):
    """
    写入数据到 Excel 文件中。如果文件存在，则新增一个 Sheet；
    如果文件不存在，则创建新文件。
    :param file_path: 文件路径
    :param value: 要写入的数据（包含 Company 的对象列表）
    """
    if not file_path:
        return

    if not value:
        return

    # 转换数据为二维数组（含表头）
    header = CellItem()  # 表头
    cell_list = [header.values()]  # 初始化数据列表
    for company in value:
        item = CellItem(company)
        cell_list.append(item.values())

    if os.path.exists(file_path):
        # 文件存在，加载工作簿
        workbook = load_workbook(file_path)

        # 根据已有工作表生成新 sheet 名称
        new_sheet_name = get_new_sheet_name(workbook.sheetnames)

        # 创建新的 Sheet
        sheet = workbook.create_sheet(new_sheet_name)
    else:
        # 文件不存在，创建新工作簿和 Sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet"

    # 将数据写入 Sheet
    for row_idx, row_data in enumerate(cell_list, start=1):  # Excel 行号从 1 开始
        for col_idx, cell_value in enumerate(row_data, start=1):  # Excel 列号从 1 开始
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # 保存文件
    workbook.save(file_path)